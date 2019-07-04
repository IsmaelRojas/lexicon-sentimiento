import openpyxl
import sys
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# -- global --
global cambio_polaridad
# -- /global --

# -- definiciones --
def leer_archivo(txt, tipo_separacion, tipo_ingreso):
    texto = open(txt, "r")
    linea_texto = texto.read()
    texto.close()
    if tipo_ingreso == "usuario":
        caracter_eliminados = ['.', ',', "'"]
        for i in range (0, len(caracter_eliminados)):
           linea_texto = linea_texto.replace(caracter_eliminados[i], '')
        linea_texto = linea_texto.replace('\n', ' ')
    return(linea_texto.split(tipo_separacion))

def leer_excel(excel):
    texto = openpyxl.load_workbook(excel)
    hoja = texto[texto.sheetnames[0]]
    lista = []
    for i in range (1, hoja.max_row):
        lista.append((str(hoja["A"+str(i)].value)) + " " + str(hoja["B"+str(i)].value))
    return lista

#usado para leer un excel con problemas de caracteres incompatibles a través de un .txt generado por Microsoft Excel
def lexicon_sentimiento(txt):
    archivo = open("lexicon-SENTIMIENTO.txt","r")
    texto_archivo = archivo.read()
    archivo.close()
    texto_archivo = texto_archivo.split("\t\t\t\t\t\t")

    for i in range(len(texto_archivo)):
        texto_archivo[i] = texto_archivo[i].replace("positivo", "positivo\n")
        texto_archivo[i] = texto_archivo[i].replace("negativo", "negativo\n")
        texto_archivo[i] = texto_archivo[i].replace("neutro", "neutro\n")
        texto_archivo[i] = texto_archivo[i].replace("?", "")
        texto_archivo[i] = texto_archivo[i].replace("\t", " ")
        texto_archivo[i] = texto_archivo[i].replace("\n", "")
    return texto_archivo

def palabras_minusculas(lista):
    for i in range (0, len(lista)-1):
        lista[i] = lista[i].lower()

#def aumento_cantidad_polaridad(polaridad, cambio_polaridad, cantidad_polaridad, intensificador):
def aumento_cantidad_polaridad(polaridad, cantidad_polaridad, intensificador):
    global cambio_polaridad
    if polaridad == "neutro" or polaridad == "neu":
        cantidad_polaridad[2] += 1
    polaridad_positiva = ["positivo", "pos"]
    polaridad_negativa = ["negativo", "neg"]
    if polaridad in polaridad_positiva:
        if cambio_polaridad:
            cantidad_polaridad[1] += (1 * intensificador)
            cambio_polaridad = False
        else:
            cantidad_polaridad[0] += (1 * intensificador)
    if polaridad in polaridad_negativa:
        if cambio_polaridad:
            cantidad_polaridad[0] += (1 * intensificador)
            cambio_polaridad = False
        else:
            cantidad_polaridad[1] += (1 * intensificador)

def texto_ingresado():
    opcion = ""
    while opcion != "1" and opcion != "2":
        print("Selecione una opción:\n1.- Ingreso de texto por teclado.\n2.- Ingreso de texto selccionando un archivo de texto.")
        opcion = input("Ingrese su opción: ")
    if opcion == "1":
        archivo = input("Ingrese el texto: ")
        caracter_eliminados = ['.', ',', "'"]
        for i in range (0, len(caracter_eliminados)):
           archivo = archivo.replace(caracter_eliminados[i], '')
        archivo = archivo.split(" ")
    else:
        Tk().withdraw()
        archivo = leer_archivo(askopenfilename(), " ", "usuario")
    return archivo
# -- /definiciones --

# -- variables -- #
print("Procesando lexicon.txt, lexicon-SENTIMIENTO.xlsx y badWords.csv\n...")
lexicon_emoticon = leer_archivo("lexicono.txt", "\n", "programador")
lexicon_sentimiento = leer_excel('lexicon-SENTIMIENTO.xlsx')
malas_palabras = leer_archivo("badWords.csv", "\n", "programador")
texto_ingresado = texto_ingresado()
palabras_minusculas(texto_ingresado)
intensificador = ["muy", "tan", "demasiado", "demasiada", "bastante", "terriblemente"]
palabra_negacion = ["no", "tampoco", "nadie", "jamas" , "ni", "sin", "nada", "nunca", "ninguno"]
#[0] polaridad positiva, [1] polaridad negativa, [2] polaridad neutra
cantidad_polaridad = [0, 0, 0]
cantidad_malas_palabras = 0
veces_intensificador = 0
aumento_intensificador = 1
cambio_polaridad = False
# -- /variables -- #

# -- principal -- #

print("Procesando texto ingresado, realizando comparación con lexicon, lexicon-SENTIMIENTO y badWords\n...")
for i in range (0, len(texto_ingresado)):

    if texto_ingresado[i] in malas_palabras:
        cantidad_malas_palabras += 1

    if veces_intensificador > 0:
        veces_intensificador -= 1
        if veces_intensificador == 0:
            aumento_intensificador = 1

    if texto_ingresado[i] in intensificador:
        veces_intensificador = 2
        aumento_intensificador = 2
        i += 1
    elif texto_ingresado[i] in palabra_negacion:
        if cambio_polaridad == False:
            cambio_polaridad = True

    for j in range(0, len(lexicon_emoticon)):
        temporal = lexicon_emoticon[j].split(" ")
        if texto_ingresado[i] == temporal[0]:
            aumento_cantidad_polaridad(temporal[1], cantidad_polaridad, aumento_intensificador)

    for k in range(0, len(lexicon_sentimiento)):
        temporal = lexicon_sentimiento[k].split(" ")
        if texto_ingresado[i] == temporal[0]:
            aumento_cantidad_polaridad(temporal[1], cantidad_polaridad, aumento_intensificador)

print("Cantidad de palabra(s) del texto: ", len(texto_ingresado))
print("Cantidad de palabra(s) positiva(s): ",cantidad_polaridad[0])
print("Cantidad de palabra(s) negativa(s): ",cantidad_polaridad[1])
print("Cantidad de palabra(s) neutra(s): ",cantidad_polaridad[2])
print("Cantidad de mala(s) palabra(s): ",cantidad_malas_palabras)
if cantidad_polaridad[0] > cantidad_polaridad[1] and cantidad_polaridad [0] > cantidad_polaridad[2]:
    print("El texto ingresado tiene polaridad positiva")
if cantidad_polaridad[1] > cantidad_polaridad[0] and cantidad_polaridad [1] > cantidad_polaridad[2]:
    print("El texto ingresado tiene polaridad negativa")
if cantidad_polaridad[2] > cantidad_polaridad[0] and cantidad_polaridad [2] > cantidad_polaridad[1]:
    print("El texto ingresado tiene polaridad neutra")
# -- /principal -- #